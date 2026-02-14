#!/usr/bin/env python3
"""
Script to create company profiles for bioplastic companies and infer missing company types.
"""

import openpyxl
import os
import re
from pathlib import Path
from datetime import datetime

def slugify_company_name(name):
    """Convert company name to filename slug."""
    if not name:
        return ""
    # Replace & with 'and'
    slug = name.replace(' & ', ' and ')
    slug = slug.replace('&', 'and')
    # Convert to lowercase
    slug = slug.lower()
    # Replace spaces and special characters with hyphens
    slug = re.sub(r'[^\w\-]', '-', slug)
    # Remove consecutive hyphens
    slug = re.sub(r'-+', '-', slug)
    # Remove leading/trailing hyphens
    slug = slug.strip('-')
    return slug

def infer_company_type(company_name, description):
    """Infer company type based on description and name."""
    if not description:
        description = ""

    desc_lower = description.lower()
    name_lower = company_name.lower()

    # Define type inference rules
    type_rules = [
        (r'\b(producer|manufacturer|produces|producing|making|makes)\b', "Bioplastic Producer"),
        (r'\b(university|research|institute|institute|academia|academic)\b', "University"),
        (r'\b(compounder|compounding|compounds|compounding)\b', "Compounder"),
        (r'\b(converter|conversion|converts|converting)\b', "Converter"),
        (r'\b(equipment|machinery|machine|equipment manufacturer)\b', "Equipment Manufacturer"),
        (r'\b(test|testing|certify|certification|standard|auditing)\b', "Testing/Certification Company"),
        (r'\b(additive|ingredient|material|resin)\b', "Additive Producer"),
        (r'\b(technology|develops|developing|develops|innovates|innovation)\b', "Technology Company"),
        (r'\b(polymer|polyester|polylactic|PLA|PHA|PBS|PBAT|biopolymer)\b', "Bioplastic Producer"),
    ]

    # Check each rule
    for pattern, company_type in type_rules:
        if re.search(pattern, desc_lower) or re.search(pattern, name_lower):
            return company_type

    # Default for ambiguous cases
    return "Technology Company"

def create_markdown_profile(company_name, company_type, country, description,
                            primary_materials, market_segments, status,
                            publicly_listed, stock_ticker, website):
    """Create markdown profile content for a company."""

    # Handle missing values
    country = country if country and str(country).strip() != '' else "Unknown"
    status = status if status and str(status).strip() != '' else "Unknown"
    website = website if website and str(website).strip() != '' else ""
    primary_materials = primary_materials if primary_materials and str(primary_materials).strip() != '' else "Not specified"
    market_segments = market_segments if market_segments and str(market_segments).strip() != '' else "Not specified"
    description = description if description and str(description).strip() != '' else "No description available."

    # Handle publicly listed
    publicly_listed_bool = False
    if publicly_listed:
        publicly_listed_str = str(publicly_listed).lower().strip()
        publicly_listed_bool = publicly_listed_str == 'yes' or publicly_listed_str == 'true'

    # Handle stock ticker
    if stock_ticker and str(stock_ticker).strip() != '' and str(stock_ticker).lower() != 'none':
        stock_ticker_str = str(stock_ticker).strip()
    else:
        stock_ticker_str = ".nan"

    # Create short description (first 100-150 chars)
    description_text = str(description)
    short_desc = description_text[:150]
    if len(short_desc) < len(description_text):
        short_desc = short_desc.rsplit(' ', 1)[0] + "..."

    # Build markdown
    markdown = f"""---
title: {company_name}
company_type: {company_type}
headquarters: {country}
status: {status}
short_description: {short_desc}
website: {website}
primary_materials: {primary_materials}
market_segments: {market_segments}
publicly_listed: {str(publicly_listed_bool).lower()}
stock_ticker: {stock_ticker_str}
date: '2026-02-13'
draft: false
---

## Overview

{description_text if description_text != "No description available." else description}

### Primary Materials

{primary_materials}

### Market Segments

{market_segments}

---

*Last updated: February 13, 2026*
"""
    return markdown

def main():
    excel_path = '/home/sven/bioplastic-website/companies.xlsx'
    content_dir = '/home/sven/bioplastic-website/content/companies'

    # Load Excel file
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Get existing profiles
    existing_profiles = set()
    if os.path.exists(content_dir):
        existing_profiles = set(f[:-3] for f in os.listdir(content_dir) if f.endswith('.md'))

    profiles_created = 0
    types_inferred = 0
    types_inferred_list = []
    errors = []

    # Process each company
    for row_idx in range(2, ws.max_row + 1):
        try:
            company_name = ws.cell(row=row_idx, column=1).value
            company_type = ws.cell(row=row_idx, column=2).value
            country = ws.cell(row=row_idx, column=3).value
            website = ws.cell(row=row_idx, column=4).value
            description = ws.cell(row=row_idx, column=5).value
            primary_materials = ws.cell(row=row_idx, column=6).value
            market_segments = ws.cell(row=row_idx, column=7).value
            status = ws.cell(row=row_idx, column=8).value
            publicly_listed = ws.cell(row=row_idx, column=9).value
            stock_ticker = ws.cell(row=row_idx, column=10).value

            if not company_name:
                continue

            # Skip companies with no meaningful data
            has_data = (country and str(country).strip() != '') or (description and str(description).strip() != '')
            if not has_data:
                continue

            # Infer missing type
            if not company_type or str(company_type).strip() == '':
                company_type = infer_company_type(company_name, description)
                ws.cell(row=row_idx, column=2).value = company_type
                types_inferred += 1
                types_inferred_list.append(f"{company_name}: {company_type}")

            # Create profile if it doesn't exist
            slug = slugify_company_name(company_name)
            if slug and slug not in existing_profiles:
                markdown_content = create_markdown_profile(
                    company_name, company_type, country, description,
                    primary_materials, market_segments, status,
                    publicly_listed, stock_ticker, website
                )

                profile_path = os.path.join(content_dir, f"{slug}.md")
                with open(profile_path, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)

                profiles_created += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    # Save updated Excel file
    try:
        wb.save(excel_path)
    except Exception as e:
        errors.append(f"Failed to save Excel file: {str(e)}")

    # Print report
    print("=" * 70)
    print("COMPANY PROFILE CREATION AND TYPE INFERENCE REPORT")
    print("=" * 70)
    print(f"\nProfiles Created: {profiles_created}")
    print(f"Company Types Inferred: {types_inferred}")
    print(f"\nTotal Companies in Excel: {ws.max_row - 1}")
    print(f"Total Company Profiles Now: {profiles_created + len(existing_profiles)}")

    if types_inferred > 0 and types_inferred_list:
        print(f"\n\nCompanies with Types Inferred ({len(types_inferred_list)}):")
        for entry in types_inferred_list[:20]:
            print(f"  - {entry}")
        if len(types_inferred_list) > 20:
            print(f"  ... and {len(types_inferred_list) - 20} more")

    if errors:
        print(f"\nErrors Encountered: {len(errors)}")
        for error in errors[:10]:  # Show first 10 errors
            print(f"  - {error}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")
    else:
        print("\nNo errors encountered!")

    print("\n" + "=" * 70)

if __name__ == "__main__":
    main()
